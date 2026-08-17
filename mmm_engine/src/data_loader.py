# -*- coding: utf-8 -*-
"""Excel data loader for MMM engine.

Features:
- Auto header-row detection
- Column name auto-mapping with confirmation table
- Auto data cleansing (currency symbols, negatives, gaps)
- Weekly / daily frequency auto-detection
- YAML config save/load for reproducible column mapping
"""
import re
import yaml
import numpy as np
import pandas as pd
from pathlib import Path


def _apply_ev_transform(arr: np.ndarray, transform: str, dates) -> np.ndarray:
    """external_vars の前処理変換を適用する。"""
    arr = arr.copy().astype(float)
    t = (transform or 'none').lower()
    if t == 'none':
        return arr
    if t == 'log1p':
        return np.log1p(np.maximum(arr, 0))
    if t == 'standardize':
        std = arr.std()
        return (arr - arr.mean()) / std if std > 0 else arr - arr.mean()
    if t == 'seasonal_deviation':
        months = pd.DatetimeIndex(dates).month
        for m in range(1, 13):
            mask = months == m
            if mask.sum() > 0:
                arr[mask] -= arr[mask].mean()
        return arr
    return arr

# ── Known channel definitions ───────────────────────────────────────────────
# Each entry: canonical_name → list of keyword fragments to match column names
_CHANNEL_KEYWORDS_RAW = {
    # ── 秤クライアント ─────────────────────────────────────────────
    'SEM_PC':        ['sem', 'pc'],
    'SEM_MOBILE':    ['sem', 'mobile'],
    'SEM_TABLET':    ['sem', 'tablet'],
    'MOVIE_PC':      ['movie', 'pc'],
    'MOVIE_MOBILE':  ['movie', 'mobile'],
    'MOVIE_TABLET':  ['movie', 'tablet'],
    'DEMAND_PC':     ['demand', 'pc'],
    'DEMAND_MOBILE': ['demand', 'mobile'],
    'DEMAND_TABLET': ['demand', 'tablet'],
    'Pmax_PC':       ['pmax', 'pc'],
    'Pmax_MOBILE':   ['pmax', 'mobile'],
    'Pmax_TABLET':   ['pmax', 'tablet'],
    'META':          ['meta', 'facebook', 'fb'],
    'X_XT_en':       ['x_xt_en', 'x xt en', 'twitter_xt_en', 'xt_en'],
    'X_XT':          ['x_xt', 'x xt', 'twitter_xt', 'xt'],
    'X_LP_en':       ['x_lp_en', 'lp_en'],
    'X_LP':          ['x_lp', 'twitter_lp', 'lp'],
    'X_MV_en':       ['x_mv_en', 'mv_en'],
    'X_MV':          ['x_mv', 'twitter_mv', 'mv'],
    'X_TDL_en':      ['x_tdl_en', 'tdl_en'],
    'X_TDL':         ['x_tdl', 'twitter_tdl', 'tdl'],
    'X_OTHER_en':    ['x_other_en', 'other_en'],
    'X_OTHER':       ['x_other', 'twitter_other', 'x other'],
    'Tver':          ['tver'],
    # ── 汎用チャネル（SNS）─────────────────────────────────────────
    'INSTAGRAM':     ['instagram', 'insta'],
    'LINE':          ['line'],
    'YOUTUBE':       ['youtube', 'yt'],
    'X':             ['x_cl', 'x_imp', 'x_s', 'x_cost'],
    'GOOGLE_DEMAND': ['google_demand', 'demand_gen', 'demandgen'],
    'YAHOO':         ['yahoo', 'yss', 'yd'],
    'TIKTOK':        ['tiktok', 'tt'],
    'TWITTER':       ['twitter', 'x_ad', 'xad'],
    'LINKEDIN':      ['linkedin'],
    'PINTEREST':     ['pinterest', 'pin'],
    'SNAPCHAT':      ['snapchat', 'snap'],
    'SPOTIFY':       ['spotify'],
    'PODCAST':       ['podcast'],
    'INFLUENCER':    ['influencer', 'infl', 'インフルエンサー'],
    # ── 汎用チャネル（検索・ディスプレイ）────────────────────────
    'GOOGLE_SEARCH': ['google_search', 'google search'],
    'GOOGLE_DISPLAY':['google_display', 'google display', 'gdn_display'],
    'GOOGLE_VIDEO':  ['google_video', 'google video', 'dv360'],
    'GOOGLE_PMAX':   ['google_pmax', 'google pmax', 'p_max'],
    'GOOGLE_SHOPPING':['google_shopping', 'google shopping', 'shopping'],
    'DISPLAY':       ['display', 'gdn', 'バナー'],
    'RETARGETING':   ['retargeting', 'remarketing', 'リターゲ', 'criteo', 'rtg'],
    'NATIVE':        ['native', 'ネイティブ'],
    'DSP':           ['dsp', 'programmatic', 'プログラマティック'],
    # ── 汎用チャネル（国内メディア）──────────────────────────────
    'SMARTNEWS':     ['smartnews', 'smart_news'],
    'GUNOSY':        ['gunosy'],
    'POPLN':         ['popln'],
    'LOGLY':         ['logly'],
    'ABEMA':         ['abema'],
    'NICONICO':      ['niconico', 'nicovideo', 'nico'],
    # ── 汎用チャネル（オフライン・マス）──────────────────────────
    'TV':            ['tv', 'television', 'テレビ', 'テレビ広告'],
    'RADIO':         ['radio', 'ラジオ'],
    'NEWSPAPER':     ['newspaper', '新聞'],
    'MAGAZINE':      ['magazine', '雑誌', 'mag'],
    'OOH':           ['ooh', 'outdoor', '屋外', '看板', 'billboard', 'transit'],
    'FLYER':         ['flyer', 'チラシ', 'leaflet'],
    'DM':            ['direct_mail', 'dm', 'ダイレクトメール'],
    # ── 汎用チャネル（デジタル施策）──────────────────────────────
    'EMAIL':         ['email', 'メール', 'mail'],
    'PUSH':          ['push', 'プッシュ通知', 'push_notif'],
    'SMS':           ['sms'],
    'SEO':           ['seo', 'organic'],
    'AFFILIATE':     ['affiliate', 'aff'],
    'CAMPAIGN':      ['campaign', 'キャンペーン'],
    'AMAZON':        ['amazon', 'amz', 'amazon_ads'],
    'PR':            ['pr', 'press'],
    'RECRUITING':    ['recruiting', 'recruit', '採用'],
    # ── CORDER クライアント ────────────────────────────────────────
    'K_BUKKA':       ['k_bukka', 'bukka'],
    'K_PLAZA':       ['k_plaza', 'plaza'],
    'KEICHO':        ['keicho_sekisan', 'keicho sekisan', 'keicho'],
    'BSIJ':          ['bsij_sekisan', 'bsij sekisan', 'bsij'],
    'SALES_AD':      ['sales_s', 'sales ad'],
    'KC_SEKISAN':    ['kc_sekisan', 'kc sekisan'],
    'KC_CONST_PRICE':['kc_const_price', 'kc const price'],
}

# ブランド表記の上書きマップ（_to_title の変換後に適用）
_CH_BRAND_NAMES: dict[str, str] = {
    'Youtube':  'YouTube',
    'Tiktok':   'TikTok',
    'Linkedin': 'LinkedIn',
    'Line':     'LINE',
    'Seo':      'SEO',
}


def _ch_title(name: str) -> str:
    """ALL_CAPSのチャネルキーをTitle_Caseに変換。Mixed_Caseはそのまま。
    例: GOOGLE_DEMAND → Google_Demand, META → Meta, YOUTUBE → YouTube
    """
    converted = '_'.join(s.capitalize() for s in name.split('_')) if name == name.upper() else name
    return _CH_BRAND_NAMES.get(converted, converted)


CHANNEL_KEYWORDS: dict = {_ch_title(k): v for k, v in _CHANNEL_KEYWORDS_RAW.items()}

# クライアント専用チャネル（UIドロップダウンから除外）
_CLIENT_SPECIFIC_KEYS: frozenset = frozenset({
    # 秤クライアント: デバイス別分割
    'SEM_PC', 'SEM_MOBILE', 'SEM_TABLET',
    'MOVIE_PC', 'MOVIE_MOBILE', 'MOVIE_TABLET',
    'DEMAND_PC', 'DEMAND_MOBILE', 'DEMAND_TABLET',
    'Pmax_PC', 'Pmax_MOBILE', 'Pmax_TABLET',
    # 秤クライアント: Xキャンペーンタイプ分割
    'X_XT_en', 'X_XT', 'X_LP_en', 'X_LP',
    'X_MV_en', 'X_MV', 'X_TDL_en', 'X_TDL', 'X_OTHER_en', 'X_OTHER',
    # CORDERクライアント専用
    'K_BUKKA', 'K_PLAZA', 'KEICHO', 'BSIJ', 'SALES_AD', 'KC_SEKISAN', 'KC_CONST_PRICE',
})

# UIドロップダウン用: 汎用チャネルのみ（クライアント専用を除外）
GENERIC_CHANNEL_KEYWORDS: dict = {
    _ch_title(k): v
    for k, v in _CHANNEL_KEYWORDS_RAW.items()
    if k not in _CLIENT_SPECIFIC_KEYS
}

# Role detection keywords (Japanese only — romaji variants are auto-generated at module load)
COST_KW   = ['cost', 'コスト', '費用', '出稿費', 'spend', '支出', '金額', 'amt', 'amount', '円', 'budget']
MEDIA_KW  = ['cl', 'click', 'クリック', 'imp', 'impression', 'インプレッション',
              '再生', 'view', 'en', 'eng', 'engagement', 'エンゲージメント',
              'pv', 'session', 'reach', 'freq', 'vtr', 'pl', 'play', 'saisei']
DATE_KW   = ['date', '日付', 'dt', '日', 'ymd', 'day']
CV_KW     = ['cv', 'コンバージョン', 'conversion', 'cvr', 'cv_uu', 'uu', 'purchase',
             '購入', '申込', '成約', '獲得', 'lead']
CONTROL_KW = ['rain', '雨', '降水', '降水量', 'snow', '雪', '積雪', '積雪量',
               'seo', 'organic', 'temperature', '気温', 'holiday', '祝日', 'repeat', 'index',
               'appt', 'appointment', 'アポ', '商談']


def _normalize_romaji(s: str) -> str:
    """Contract long vowels for loose romaji matching: ou→o, oo→o, uu→u."""
    s = re.sub(r'ou', 'o', s)
    s = re.sub(r'oo', 'o', s)
    s = re.sub(r'uu', 'u', s)
    return s


def _build_romaji_role_map() -> dict[str, str]:
    """Generate normalized romaji → role mapping from Japanese keywords in KW lists.

    Called once at module load. Requires pykakasi; returns {} gracefully if absent.
    Only entries with normalized romaji length ≥ 4 are stored to avoid false positives.
    """
    try:
        import pykakasi as _pk
        kks = _pk.kakasi()
    except ImportError:
        return {}
    role_map: dict[str, str] = {}
    for kw_list, role in [(COST_KW, 'cost'), (MEDIA_KW, 'media'),
                           (CV_KW, 'cv'), (CONTROL_KW, 'control')]:
        for kw in kw_list:
            if not any(ord(ch) > 127 for ch in kw):
                continue
            result = kks.convert(kw)
            for variant_key in ('hepburn', 'passport'):
                raw = ''.join(item.get(variant_key, '') for item in result)
                if not raw:
                    continue
                normalized = _normalize_romaji(raw)
                if len(normalized) >= 4 and normalized.isascii():
                    role_map.setdefault(normalized, role)
    return role_map


_ROMAJI_ROLE_MAP: dict[str, str] = _build_romaji_role_map()


# ── Header row auto-detection ────────────────────────────────────────────────

def detect_header_row(excel_path: str, sheet_name: str = None, max_scan: int = 15) -> int:
    """Scan first max_scan rows; return 1-indexed row number that looks like a header.

    Criteria: row where a DATE-like cell AND CV-like cell can be found in the same row.
    Fallback: first row that has ≥ 5 string cells.
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True, keep_vba=False)
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    best_row = 1
    best_score = 0

    max_row = ws.max_row or 10
    max_col = ws.max_column or 50
    for row_idx in range(1, min(max_scan + 1, max_row + 1)):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, min(max_col + 1, 80))]
        str_vals = [str(v).lower().strip() for v in vals if v is not None and isinstance(v, str)]

        n_str = len(str_vals)
        has_date = any(any(kw in s for kw in DATE_KW) for s in str_vals)
        has_cv   = any(any(kw in s for kw in CV_KW)   for s in str_vals)
        has_ch   = any(any(kw in s for kw in ['sem', 'pmax', 'meta', 'movie', 'demand', 'tver'])
                       for s in str_vals)

        score = n_str + has_date * 5 + has_cv * 5 + has_ch * 3
        if score > best_score:
            best_score = score
            best_row = row_idx

    wb.close()
    return best_row


def detect_sheet_name(excel_path: str) -> str:
    """Return first sheet name (openpyxl)."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True, keep_vba=False)
    name = wb.sheetnames[0]
    wb.close()
    return name


# ── Column auto-mapping ──────────────────────────────────────────────────────

def _col_key(col_name: str) -> str:
    """Normalize column name for matching."""
    s = col_name.lower()
    s = re.sub(r'[_\-\s]+', '_', s)
    s = s.strip('_')
    return s


def _score_channel(col: str, channel: str, kws: list) -> float:
    """Score how well col matches channel given its keyword list."""
    c = _col_key(col)
    ch_norm = _col_key(channel)

    # Hard exclusion: if channel does not end with _en, penalize _en columns heavily
    ch_is_en  = ch_norm.endswith('_en')
    col_has_en = c.endswith('_en') or '_en_' in c or c.endswith('en')

    if ch_is_en and not col_has_en:
        return 0.0   # _en channel must use _en column
    if not ch_is_en and col_has_en:
        return 0.0   # non-_en channel must NOT use _en column

    score = 0.0
    for kw in kws:
        if kw in c:
            score += 1.0 / len(kws)
    # Bonus for exact channel name match
    if c.startswith(ch_norm) or ch_norm == c or c == ch_norm + '_cl' or c == ch_norm + '_cost':
        score += 0.8
    elif ch_norm in c:
        score += 0.5
    return score


def _detect_role(col: str) -> str:
    """Detect column role from name keywords, with romaji fallback for ASCII column names."""
    c = _col_key(col)
    if any(kw in c for kw in [_col_key(k) for k in COST_KW]):
        return 'cost'
    # _s suffix convention (e.g. google_search_s, facebook_s) = spend/cost
    if c.endswith('_s') and not any(kw in c for kw in [_col_key(k) for k in MEDIA_KW]):
        return 'cost'
    if any(kw in c for kw in [_col_key(k) for k in DATE_KW]):
        return 'date'
    if any(kw in c for kw in [_col_key(k) for k in CV_KW]):
        return 'cv'
    if any(kw in c for kw in [_col_key(k) for k in CONTROL_KW]):
        return 'control'
    if any(kw in c for kw in [_col_key(k) for k in MEDIA_KW]):
        return 'media'
    # Romaji fallback: ASCII-only column names may be romanized Japanese.
    # Normalize long vowels (ou→o etc.) then check auto-generated romaji map.
    if c.isascii() and len(c) >= 4 and _ROMAJI_ROLE_MAP:
        c_norm = _normalize_romaji(c)
        for romaji, role in _ROMAJI_ROLE_MAP.items():
            if romaji in c_norm:
                return role
    return 'unknown'


def _detect_role_from_values(series: pd.Series) -> str | None:
    """Detect role override from actual column values.

    Returns a role string to override the name-based detection, or None if no override.
    Only downgrades media/cost/unknown → control; never upgrades.

    Rules (in priority order):
      bool / {0,1}   → control  (binary flag)
      null-or-1 flag → control  (presence flag: 0=absent, 1=present)
      all 0.0–1.0 float → control  (rate / ratio, e.g. CTR, CVR)
      constant column   → control  (zero variance, useless for regression)
    """
    s = series.dropna()
    if len(s) < 3:
        return None

    # Try numeric conversion
    try:
        s_num = pd.to_numeric(s, errors='raise')
    except (ValueError, TypeError):
        return None  # non-numeric column — leave as-is

    uniq = set(s_num.unique())

    # Bool / {0, 1} binary flag
    if uniq <= {0.0, 1.0}:
        return 'control'

    # Null-or-1 presence flag (non-zero values are exclusively 1)
    non_zero = s_num[s_num != 0]
    if len(non_zero) > 0 and set(non_zero.unique()) == {1.0}:
        return 'control'

    # Rate / ratio: float values strictly in [0, 1]
    if s_num.max() <= 1.0 and s_num.min() >= 0.0 and s_num.dtype.kind == 'f':
        return 'control'

    # Constant column (zero or near-zero variance)
    if s_num.std() < 1e-9:
        return 'control'

    return None


def _effective_role(col: str, series: pd.Series | None = None) -> str:
    """Combine name-based and value-based role detection.

    Value-based heuristics apply ONLY to 'unknown' columns — those whose
    name contains no recognizable keyword. Columns already identified as
    cost, media, date, cv, or control by name are trusted as-is.

    This prevents mis-classifying sparse click columns (which briefly look
    like {0,1} binary flags) or ratio-like viewership columns.
    """
    name_role = _detect_role(col)
    if series is None or name_role != 'unknown':
        return name_role
    val_override = _detect_role_from_values(series)
    return val_override if val_override is not None else name_role


def auto_map_columns(df: pd.DataFrame) -> dict:
    """Auto-map DataFrame columns to MMM roles.

    Returns dict with keys:
        date_col, cv_col,
        channel_map: {ch_name: {'media': col, 'cost': col}},
        control_cols: [col, ...],
        unmapped: [col, ...]
    """
    cols = [c for c in df.columns if c and str(c).strip()]

    # Precompute effective role per column (name + value heuristics)
    has_values = len(df) >= 10
    _role_cache: dict = {}
    for c in cols:
        series = df[c] if has_values else None
        _role_cache[c] = _effective_role(str(c), series)

    # Step 1: Find date and CV columns
    date_col = None
    cv_col   = None
    for c in cols:
        role = _role_cache[c]
        if role == 'date' and date_col is None:
            date_col = c
        if role == 'cv' and cv_col is None:
            cv_col = c

    # Step 2: Score each column against each channel
    # Separate cost / media / control pools using effective role
    cost_cols    = {c for c in cols if _role_cache[c] == 'cost'}
    media_cols   = {c for c in cols if _role_cache[c] in ('media', 'unknown')
                    and c not in (date_col, cv_col)}
    control_cols = [c for c in cols if _role_cache[c] == 'control']

    channel_map = {}
    for ch, kws in CHANNEL_KEYWORDS.items():
        # Find best cost col
        best_cost = None
        best_cost_score = 0.0
        for c in cost_cols:
            s = _score_channel(str(c), ch, kws)
            if s > best_cost_score:
                best_cost_score = s
                best_cost = c

        # Find best media col (exclude cost cols)
        best_media = None
        best_media_score = 0.0
        for c in media_cols:
            if c == best_cost:
                continue
            s = _score_channel(str(c), ch, kws)
            if s > best_media_score:
                best_media_score = s
                best_media = c

        if best_media or best_cost:
            channel_map[ch] = {
                'media':       best_media,
                'cost':        best_cost,
                'media_score': round(best_media_score, 2),
                'cost_score':  round(best_cost_score,  2),
            }

    mapped_cols = {date_col, cv_col}
    for ch, m in channel_map.items():
        if m['media']: mapped_cols.add(m['media'])
        if m['cost']:  mapped_cols.add(m['cost'])
    mapped_cols.update(control_cols)
    mapped_cols.discard(None)

    unmapped = [c for c in cols if c not in mapped_cols]

    return {
        'date_col':    date_col,
        'cv_col':      cv_col,
        'channel_map': channel_map,
        'control_cols': control_cols,
        'unmapped':    unmapped,
    }


def print_mapping_table(mapping: dict, freq: str = 'daily') -> str:
    """Return a human-readable mapping confirmation table string."""
    lines = []
    lines.append('=' * 70)
    lines.append(f'  MMM列マッピング確認（頻度: {freq}）')
    lines.append('=' * 70)
    lines.append(f'  DATE列  : {mapping["date_col"]}')
    lines.append(f'  CV列    : {mapping["cv_col"]}')
    lines.append(f'  コントロール変数: {mapping["control_cols"]}')
    lines.append('-' * 70)
    lines.append(f'  {"チャネル":<20} {"メディア列":<30} {"コスト列":<25}')
    lines.append('-' * 70)
    for ch, m in mapping['channel_map'].items():
        media = m['media'] or '（未検出）'
        cost  = m['cost']  or '（未検出）'
        warn  = ' ⚠' if not m['media'] and not m['cost'] else ''
        warn  += ' △媒体未検出' if not m['media'] and m['cost'] else ''
        warn  += ' △コスト未検出' if m['media'] and not m['cost'] else ''
        lines.append(f'  {ch:<20} {media:<30} {cost:<25}{warn}')
    if mapping['unmapped']:
        lines.append('-' * 70)
        lines.append(f'  未マッピング列: {mapping["unmapped"]}')
    lines.append('=' * 70)
    return '\n'.join(lines)


# ── Frequency detection ──────────────────────────────────────────────────────

def detect_frequency(dates: pd.Series) -> str:
    """Detect data frequency: 'daily' or 'weekly'."""
    dates_sorted = pd.to_datetime(dates).sort_values().dropna()
    if len(dates_sorted) < 2:
        return 'daily'
    diffs = dates_sorted.diff().dropna().dt.days
    median_gap = diffs.median()
    if median_gap >= 6:
        return 'weekly'
    return 'daily'


# ── Data cleansing ───────────────────────────────────────────────────────────

def _clean_cell(val) -> float:
    """Convert a potentially messy cell value to float."""
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return max(float(val), 0.0)
    s = str(val)
    # Remove currency symbols, commas, spaces, 円
    s = re.sub(r'[¥\$€£,\s円￥]', '', s)
    # Remove trailing units like '万' → multiply by 10000
    man = '万' in s
    s = s.replace('万', '').replace('千', '').strip()
    try:
        v = float(s)
        if man:
            v *= 10000
        return max(v, 0.0)
    except ValueError:
        return 0.0


def _fill_date_gaps(df: pd.DataFrame, date_col: str, freq: str) -> pd.DataFrame:
    """Fill gaps in the date series with 0-valued rows."""
    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col])
    df = df.set_index(date_col).sort_index()
    pd_freq = 'D' if freq == 'daily' else 'W-SAT'
    full_idx = pd.date_range(df.index.min(), df.index.max(), freq='D' if freq == 'daily' else 'W')
    df = df.reindex(full_idx, fill_value=0)
    df.index.name = date_col
    return df.reset_index()


# ── Main loader ──────────────────────────────────────────────────────────────

def load_data(excel_path: str,
              sheet_name: str = None,
              header_row: int = None,
              mapping_override: dict = None,
              auto_cleanse: bool = True,
              verbose: bool = True) -> dict:
    """Load and prepare MMM data from Excel.

    Parameters
    ----------
    excel_path      : path to .xlsm or .xlsx
    sheet_name      : if None, auto-detect (first sheet)
    header_row      : 1-indexed; if None, auto-detect
    mapping_override: supply a custom channel_map dict to skip auto-mapping
    auto_cleanse    : strip ¥/commas, clip negatives, fill date gaps
    verbose         : print mapping table

    Returns
    -------
    dict with keys:
        dates, cv_uu, media, costs, controls, n_days, freq,
        mapping, raw_df
    """
    path = Path(excel_path)

    # Auto-detect sheet
    if sheet_name is None:
        sheet_name = detect_sheet_name(str(path))

    # Auto-detect header row
    if header_row is None:
        header_row = detect_header_row(str(path), sheet_name)

    if verbose:
        print(f'  シート: {sheet_name}  ヘッダー行: {header_row}行目')

    # Read raw
    df = pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=header_row - 1,
        engine='openpyxl',
    )
    # Drop completely empty rows
    df = df.dropna(how='all').reset_index(drop=True)

    # Get mapping
    mapping = mapping_override or auto_map_columns(df)

    date_col    = mapping['date_col']
    cv_col      = mapping['cv_col']
    channel_map = mapping['channel_map']
    control_cols = mapping['control_cols']

    if date_col is None:
        raise ValueError('DATE列が見つかりません。列名を確認してください。')
    if cv_col is None:
        raise ValueError('CV（目的変数）列が見つかりません。列名を確認してください。')

    # Parse dates; drop rows without valid date
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df = df.dropna(subset=[date_col]).reset_index(drop=True)
    df = df.sort_values(date_col).reset_index(drop=True)

    # Detect frequency
    freq = detect_frequency(df[date_col])

    if verbose:
        print(print_mapping_table(mapping, freq))
        print(f'  データ頻度: {freq} ({len(df)}レコード)')

    # Cleanse
    def get_series(col):
        if col is None or col not in df.columns:
            return np.zeros(len(df))
        if auto_cleanse:
            return np.array([_clean_cell(v) for v in df[col]], dtype=float)
        return df[col].fillna(0).astype(float).values

    cv_uu = get_series(cv_col)

    media = {}
    costs = {}
    for ch, m in channel_map.items():
        # If no media column, fall back to cost column as the signal basis
        media_col = m.get('media') or m.get('cost')
        media[ch] = get_series(media_col)
        costs[ch] = get_series(m.get('cost'))

    controls = {}
    for col in control_cols:
        controls[col] = get_series(col)

    # external_vars: YAMLで指定された外部変数を変換してcontrolsに追加
    _ev_dates = pd.DatetimeIndex(df[date_col].values)
    for ev in mapping.get('external_vars') or []:
        ev_col  = ev.get('col')
        ev_name = ev.get('name') or ev_col
        ev_tr   = ev.get('transform', 'none')
        if not ev_col:
            continue
        arr = get_series(ev_col)
        arr = _apply_ev_transform(arr, ev_tr, _ev_dates)
        controls[ev_name] = arr
        if verbose:
            print(f'  外部変数: {ev_name} (col={ev_col}, transform={ev_tr})')

    dates = df[date_col].values

    # Fill date gaps (insert 0-rows for missing dates)
    if auto_cleanse and len(dates) > 1:
        expected = pd.date_range(
            pd.Timestamp(dates[0]), pd.Timestamp(dates[-1]),
            freq='D' if freq == 'daily' else 'W'
        )
        if len(expected) > len(dates):
            gap_count = len(expected) - len(dates)
            if verbose:
                print(f'  ⚠ 日付ギャップ {gap_count}件を0で補完しました')
            # Rebuild with gap filling
            tmp = df[[date_col]].copy()
            tmp['__idx'] = range(len(df))
            tmp = tmp.set_index(date_col).reindex(expected, fill_value=np.nan).reset_index()
            new_idx = tmp['__idx'].values
            dates = expected.values

            def reindex_arr(arr):
                new = np.zeros(len(expected))
                valid = ~np.isnan(new_idx.astype(float))
                orig = new_idx[valid].astype(int)
                new[valid] = arr[orig]
                return new

            cv_uu = reindex_arr(cv_uu)
            for ch in media:
                media[ch] = reindex_arr(media[ch])
                costs[ch] = reindex_arr(costs[ch])
            for k in controls:
                controls[k] = reindex_arr(controls[k])

    n_days = len(dates)
    total_cv = int(cv_uu.sum())
    total_spend = sum(costs[ch].sum() for ch in costs)

    if verbose:
        print(f'  CV総数: {total_cv:,}件  総スペンド: {total_spend/10000:.0f}万円')

    return {
        'dates':     dates,
        'cv_uu':     cv_uu,
        'media':     media,
        'costs':     costs,
        'controls':  controls,
        'n_days':    n_days,
        'freq':      freq,
        'mapping':   mapping,
        'raw_df':    df,
    }


# ── Google Sheets loader ─────────────────────────────────────────────────────

_DEFAULT_CREDS = r'G:\マイドライブ\00_ai_company\00_admin\google-service-account\credentials.json'


def _extract_sheet_id(id_or_url: str) -> str:
    """Extract spreadsheet ID from a URL or return the ID as-is."""
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', id_or_url)
    return m.group(1) if m else id_or_url


def _open_gspread(spreadsheet_id: str, credentials_path: str = None):
    """Authorize gspread with service account and open the spreadsheet."""
    import gspread
    from google.oauth2.service_account import Credentials
    creds = Credentials.from_service_account_file(
        credentials_path or _DEFAULT_CREDS,
        scopes=['https://www.googleapis.com/auth/spreadsheets.readonly'],
    )
    return gspread.authorize(creds).open_by_key(spreadsheet_id)


def _gspread_ws_to_df(ws) -> pd.DataFrame:
    """Convert a gspread worksheet to a DataFrame (first row as header)."""
    all_values = ws.get_all_values()
    if not all_values:
        raise ValueError('シートにデータがありません。')
    return pd.DataFrame(all_values[1:], columns=all_values[0])


def load_from_sheets(
    spreadsheet_id_or_url: str,
    sheet_name: str = None,
    credentials_path: str = None,
    mapping_override: dict = None,
    auto_cleanse: bool = True,
    verbose: bool = True,
) -> dict:
    """Load and prepare MMM data from Google Sheets via service account.

    Parameters
    ----------
    spreadsheet_id_or_url : spreadsheet ID or full Google Sheets URL
    sheet_name            : worksheet name; if None, uses the first sheet
    credentials_path      : path to service account JSON; defaults to project default
    mapping_override      : custom channel_map dict to skip auto-mapping
    auto_cleanse          : strip ¥/commas, clip negatives, fill date gaps
    verbose               : print mapping table

    Returns
    -------
    Same dict structure as load_data():
        dates, cv_uu, media, costs, controls, n_days, freq, mapping, raw_df
    """
    sh = _open_gspread(_extract_sheet_id(spreadsheet_id_or_url), credentials_path)

    if sheet_name is None:
        ws = sh.sheet1
        sheet_name = ws.title
    else:
        ws = sh.worksheet(sheet_name)

    if verbose:
        print(f'  スプレッドシート: {sh.title}  シート: {sheet_name}')

    df = _gspread_ws_to_df(ws)
    df = df.replace('', np.nan).dropna(how='all').reset_index(drop=True)

    mapping     = mapping_override or auto_map_columns(df)
    date_col    = mapping['date_col']
    cv_col      = mapping['cv_col']
    channel_map = mapping['channel_map']
    control_cols = mapping['control_cols']

    if date_col is None:
        raise ValueError('DATE列が見つかりません。列名を確認してください。')
    if cv_col is None:
        raise ValueError('CV（目的変数）列が見つかりません。列名を確認してください。')

    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df = df.dropna(subset=[date_col]).reset_index(drop=True)
    df = df.sort_values(date_col).reset_index(drop=True)

    freq = detect_frequency(df[date_col])

    if verbose:
        print(print_mapping_table(mapping, freq))
        print(f'  データ頻度: {freq} ({len(df)}レコード)')

    def get_series(col):
        if col is None or col not in df.columns:
            return np.zeros(len(df))
        if auto_cleanse:
            return np.array([_clean_cell(v) for v in df[col]], dtype=float)
        return df[col].fillna(0).astype(float).values

    cv_uu = get_series(cv_col)

    media = {}
    costs = {}
    for ch, m in channel_map.items():
        media_col  = m.get('media') or m.get('cost')
        media[ch]  = get_series(media_col)
        costs[ch]  = get_series(m.get('cost'))

    controls = {}
    for col in control_cols:
        controls[col] = get_series(col)

    # external_vars: YAMLで指定された外部変数を変換してcontrolsに追加
    _ev_dates = pd.DatetimeIndex(df[date_col].values)
    for ev in mapping.get('external_vars') or []:
        ev_col  = ev.get('col')
        ev_name = ev.get('name') or ev_col
        ev_tr   = ev.get('transform', 'none')
        if not ev_col:
            continue
        arr = get_series(ev_col)
        arr = _apply_ev_transform(arr, ev_tr, _ev_dates)
        controls[ev_name] = arr
        if verbose:
            print(f'  外部変数: {ev_name} (col={ev_col}, transform={ev_tr})')

    dates = df[date_col].values

    if auto_cleanse and len(dates) > 1:
        expected = pd.date_range(
            pd.Timestamp(dates[0]), pd.Timestamp(dates[-1]),
            freq='D' if freq == 'daily' else 'W',
        )
        if len(expected) > len(dates):
            gap_count = len(expected) - len(dates)
            if verbose:
                print(f'  ⚠ 日付ギャップ {gap_count}件を0で補完しました')
            tmp = df[[date_col]].copy()
            tmp['__idx'] = range(len(df))
            tmp = tmp.set_index(date_col).reindex(expected, fill_value=np.nan).reset_index()
            new_idx = tmp['__idx'].values
            dates   = expected.values

            def reindex_arr(arr):
                out   = np.zeros(len(expected))
                valid = ~np.isnan(new_idx.astype(float))
                orig  = new_idx[valid].astype(int)
                out[valid] = arr[orig]
                return out

            cv_uu = reindex_arr(cv_uu)
            for ch in media:
                media[ch] = reindex_arr(media[ch])
                costs[ch] = reindex_arr(costs[ch])
            for k in controls:
                controls[k] = reindex_arr(controls[k])

    n_days      = len(dates)
    total_cv    = int(cv_uu.sum())
    total_spend = sum(costs[ch].sum() for ch in costs)

    if verbose:
        print(f'  CV総数: {total_cv:,}件  総スペンド: {total_spend/10000:.0f}万円')

    return {
        'dates':    dates,
        'cv_uu':    cv_uu,
        'media':    media,
        'costs':    costs,
        'controls': controls,
        'n_days':   n_days,
        'freq':     freq,
        'mapping':  mapping,
        'raw_df':   df,
    }


def detect_only_from_sheets(
    spreadsheet_id_or_url: str,
    sheet_name: str = None,
    credentials_path: str = None,
) -> dict:
    """Run column-mapping detection only on a Google Sheet (no full data load)."""
    sh = _open_gspread(_extract_sheet_id(spreadsheet_id_or_url), credentials_path)

    if sheet_name is None:
        ws = sh.sheet1
        sheet_name = ws.title
    else:
        ws = sh.worksheet(sheet_name)

    all_values = ws.get_all_values()
    if not all_values:
        raise ValueError('シートにデータがありません。')

    df_head = pd.DataFrame(all_values[1:101], columns=all_values[0]).replace('', np.nan)
    mapping = auto_map_columns(df_head)
    n       = len(all_values) - 1

    # Estimate freq from actual date gaps (fall back to row count if no date column)
    date_col_guess = mapping.get('date_col')
    if date_col_guess and date_col_guess in df_head.columns:
        freq = detect_frequency(df_head[date_col_guess])
    else:
        freq = 'weekly' if n < 400 else 'daily'

    return {
        'sheet_name': sheet_name,
        'mapping':    mapping,
        'freq_guess': freq,
        'n_rows':     n,
        'table':      print_mapping_table(mapping, freq),
    }


# ── Detect-only mode ────────────────────────────────────────────────────────

def detect_only(excel_path: str, sheet_name: str = None, header_row: int = None) -> dict:
    """Run detection only; return mapping dict for user confirmation.

    Does NOT load full data. Fast and lightweight.
    """
    path = Path(excel_path)
    if sheet_name is None:
        sheet_name = detect_sheet_name(str(path))
    if header_row is None:
        header_row = detect_header_row(str(path), sheet_name)

    # Load enough rows for value-based heuristics (binary flag, rate, constant detection)
    df_head = pd.read_excel(
        path, sheet_name=sheet_name,
        header=header_row - 1, engine='openpyxl', nrows=100,
    )
    mapping = auto_map_columns(df_head)

    # Quick frequency check — count actual data rows via pandas (openpyxl read_only may return None for max_row)
    df_full = pd.read_excel(path, sheet_name=sheet_name, header=header_row - 1, engine='openpyxl', usecols=[0])
    n = len(df_full.dropna())

    # Estimate freq from row count
    freq = 'weekly' if n < 200 else 'daily'

    return {
        'sheet_name': sheet_name,
        'header_row': header_row,
        'mapping':    mapping,
        'freq_guess': freq,
        'n_rows':     n,
        'table':      print_mapping_table(mapping, freq),
    }


# ── YAML config save / load ─────────────────────────────────────────────────

def save_mapping_yaml(mapping: dict, path: str, client: str = '') -> None:
    """Save auto-detected column mapping as a human-editable YAML config.

    The saved file can be reviewed and edited, then passed back via --config.
    """
    ch_map = mapping.get('channel_map', {})
    channels_out = {}
    for ch, m in ch_map.items():
        channels_out[ch] = {
            'spend': m.get('cost') or None,
            'media': m.get('media') or None,
        }

    doc = {
        'client':   client or '',
        'date':     mapping.get('date_col'),
        'cv':       mapping.get('cv_col'),
        'channels': channels_out,
        'controls': list(mapping.get('control_cols', [])),
    }

    header = (
        '# MMM column mapping config\n'
        '# Generated by --detect-only --save-config\n'
        '#\n'
        '# channels.<NAME>.spend  : cost/spend column (required)\n'
        '# channels.<NAME>.media  : media signal column for Adstock+Hill (null = use spend)\n'
        '#                          Set to null or remove to fall back to spend.\n'
        '#                          Extra columns (imp, etc.) should go in controls.\n'
        '# controls               : list of control/external variable columns\n'
        '#\n'
    )

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(header)
        yaml.dump(doc, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    print(f'  YAML config saved → {path}')


def load_mapping_yaml(path: str) -> dict:
    """Load a YAML config and return a mapping dict compatible with auto_map_columns output."""
    with open(path, encoding='utf-8') as f:
        doc = yaml.safe_load(f)

    ch_map = {}
    for ch, m in (doc.get('channels') or {}).items():
        spend_col = m.get('spend') if m else None
        media_col = m.get('media') if m else None
        ch_map[ch] = {
            'media':       media_col,
            'cost':        spend_col,
            'media_score': 1.0,
            'cost_score':  1.0,
        }

    mapping = {
        'date_col':      doc.get('date'),
        'cv_col':        doc.get('cv'),
        'channel_map':   ch_map,
        'control_cols':  list(doc.get('controls') or []),
        'external_vars': list(doc.get('external_vars') or []),
        'unmapped':      [],
    }
    return mapping
